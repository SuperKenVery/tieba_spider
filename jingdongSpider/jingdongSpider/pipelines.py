
import os
from openpyxl import Workbook,load_workbook
import time
"""
1.open_spider方法中，链接mongodb数据库，获取要操作的集合
2.process_item  方法中，像mongodb中插入类别数据
3.close_spider 方法中，关闭mongodb的链接
"""

class ProductPipeline(object):
    path = 'D:\\spiderFile\\dataset\\实体识别\\'+'京东商品all'+time.strftime('%Y.%m.%d.%H.%M.%S',time.localtime(time.time()))+'.xlsx'
    def __init__(self):
        # 如果已存在，则打开，否则新建
        if(os.path.exists(self.path)):
            self.wb = load_workbook(self.path)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            # self.ws.append([ '商品品牌','商品名称','商品编号','商品毛重','cpu', '运行内存','机身存储' , '存储卡'
            # , '后置摄像头','前置摄像头','屏幕大小','屏幕百分比','屏幕类型', '电池容量','颜色', '操作系统'])
            self.ws.append([ '商品品牌','商品名称','商品编号','商品毛重','处理器','机身存储' ,'屏幕大小','颜色', '操作系统','分类','商品类别'])
 
    def process_item(self, item, spider):
        line = [item['brand'],item['name'], item['code'], item['weight'], item['cpu']
        , item['file_save'] , item['screen_size'], item['color'], item['system'], item['classification'], item['product_type']]
        self.ws.append(line)
        self.wb.save(self.path)
        return item

