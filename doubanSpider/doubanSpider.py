import requests
from lxml import etree
import time
import random
from openpyxl import Workbook
import re
from wordcloud import WordCloud
import PIL .Image as image
import numpy as np
import jieba
import os
from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt

class doubanSpider:
    
    target_url = 'https://movie.douban.com/subject/26100958/comments?start={}&limit=20&sort=new_score&status=P'

    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36",
        'Cookie': 'll="118281"; bid=1E8tHh1UO7k; __utma=30149280.787827060.1593838175.1593838175.1593838175.1; __utmc=30149280; __utmz=30149280.1593838175.1.1.utmcsr=so.com|utmccn=(referral)|utmcmd=referral|utmcct=/link; ap_v=0,6.0; _vwo_uuid_v2=DFE5584FB8092E19E1C48ACB6A8C99E62|d5d4f0c4ca4c47a6ddcacacff97040ad; __gads=ID=5490f395fcb95985:T=1593838190:S=ALNI_Mbd_y4lD5XgT1pqnwj9gyQQasX2Nw; dbcl2="218965771:ytN/j1jGo58"; ck=7U_Q; __guid=236236167.3893834060458141000.1593840219409.0322; _pk_ref.100001.8cb4=%5B%22%22%2C%22%22%2C1593840220%2C%22https%3A%2F%2Faccounts.douban.com%2Faccounts%2Fpassport%2Fregister%22%5D; _pk_ses.100001.8cb4=*; push_noty_num=0; push_doumail_num=0; __utmt=1; __utmv=30149280.21896; __yadk_uid=5q5tgoXkHZk2p7qqUcXhzcqZF8yK4kpa; monitor_count=4; _pk_id.100001.8cb4=a34ccb6950d8365b.1593840220.1.1593840306.1593840220.; __utmb=30149280.9.10.1593838175'
    }

    def start_spider(self):
        result_list = []
        for i in range(0,50):
            start = i
            reponse = requests.get(self.target_url.format(start),headers=self.headers)
            # print(reponse.text)
            html = etree.HTML(str(reponse.content,'utf-8'))
            # 短评列表
            short_list = html.xpath('//div[@id="comments"]/div[@class="comment-item"]//span[@class="short"]/text()')
            print(short_list)

            times = html.xpath('//div[@class="comment-item"]//span[@class="comment-info"]/span[2]/@class')

            complte_times = html.xpath('//div[@class="comment-item"]//span[@class="comment-info"]/span[3]/@title')

            votes = html.xpath('//div[@class="comment-item"]//div[@class="comment"]/h3/span[@class="comment-vote"]/span[@class="votes"]/text()') # 赞同量

            # 用户链接列表
            user_list = html.xpath('//div[@id="comments"]/div[@class="comment-item"]//span[@class="comment-info"]/a/@href')
            for i in range(len(user_list)):
                url = user_list[i]
                item = {'short':self.clear_character_chinese(str(short_list[i]))}
                reponse = requests.get(url,headers=self.headers)
                html = etree.HTML(reponse.text)
                city = html.xpath('//div[@class="user-info"]/a/text()')
                join_date = html.xpath('//div[@class="user-info"]/div[@class="pl"]/text()')
                if(city != None):
                    if(len(city) > 0):
                        item['city'] = self.clear_character_chinese(city[0])
                    else:
                        continue
                if(join_date != None):
                    if(len(join_date)>1):
                        item['join_date'] = self.clear_character_chinese(join_date[1]).replace("加入","")
                    elif(len(join_date)>0):
                        item['join_date'] = self.clear_character_chinese(join_date[0]).replace("加入","")
                    else:
                        continue
                user_name = html.xpath('//div[@class="info"]/h1/text()')
                if(user_name != None):
                    if(len(user_name) > 0):
                        item['userName'] = self.clear_character_chinese(user_name[0])
                    else:
                        continue

                timeStr = times[i]
                if(timeStr == 'comment-time'):
                    continue
                    # item['time'] = '暂无评分'
                else:
                    item['time'] = timeStr.replace('allstar','').replace(' rating','')

                item['vote'] = str(votes[i])

                item['complete_time'] = str(complte_times[i]).split(" ")[0]
                item['detail_time'] = str(complte_times[i]).split(" ")[1]
                result_list.append(item)

                time.sleep(random.uniform(0.3,1))
        self.saveToCsv(result_list)


    def clear_character_chinese(self,sentence):
        pattern1='[a-zA-Z0-9]'
        pattern2 = '\[.*?\]'
        pattern3 = re.compile(u'[^\s1234567890:：' + '\u4e00-\u9fa5]+')
        pattern4='[’!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~]+'
        line2=re.sub(pattern2,'',sentence)   #去除表情
        new_sentence=''.join(line2.split()) #去除空白
        return new_sentence
            
    # # 保存数据到excel文件
    def saveToCsv(self,data):
        print(data)
        wb = Workbook()
        ws = wb.active
        ws.append(['短评内容','评分','赞同量','评价日期','评价时间', '用户名', '常住地址','注册时间'])
        for item in data:
            line = [item['short'], item['time'],item['vote'],item['complete_time'],item['detail_time'], item['userName'],item['city'],item['join_date']]
            ws.append(line)
            wb.save('douban.xlsx')

    # 读取短评内容
    def read_short_data(self,word_type):
        data = []
        workbook1=load_workbook('douban.xlsx')
        sheet=workbook1.get_sheet_by_name("Sheet")
        count = 0
        for row in sheet.iter_rows():
            if(count == 0):
                count = 1
                continue
            print(row[0].value)
            short = row[0].value
            short_type = row[1].value
            if (word_type == 1):
                if (int(short_type)<40):
                    continue
            elif(word_type == 2):
                if (int(short_type)>=40 or int(short_type)<=20):
                    continue
            elif(word_type == 3):
                if (int(short_type)>20):
                        continue
            short = self.clean_stopwords(short)
            data.append(short)
        return ";".join(data)

    def generWord(self,word_type):
        # 查询数据
        content = self.read_short_data(word_type)
        msg = "全部"
        if(word_type == 1):
            msg = "好评"
        elif(word_type == 2):
            msg = "中评"
        elif(word_type == 3):
            msg = "差评"
        self.get_image(content,"douban_{}.png".format(msg))

    # 生成词云
    def get_image(self,data,savePath):
        text  = self.trans_CN(data)
        wordcloud = WordCloud(
            background_color="white",
            font_path = "C:\\Windows\\Fonts\\msyh.ttc"
        ).generate(text)
        # image_produce = wordcloud.to_image()
        # image_produce.show()
        wordcloud.to_file(savePath)

    def trans_CN(self,text):
        word_list = jieba.cut(text)
        # 分词后在单独个体之间加上空格
        result = " ".join(word_list)
        return result
    
    # 去掉停用词
    def clean_stopwords(self,sentence):
        contents_list=[]
        file_dir = os.path.split(os.path.realpath(__file__))[0] + os.sep+'stop_words.txt'
        stopwords = {}.fromkeys([line.rstrip() for line in open(file_dir, encoding="utf-8")]) #读取停用词表
        stopwords_list = set(stopwords)
        words_list = jieba.lcut(sentence)
        words = [w for w in words_list if w not in stopwords_list]
        sentence=''.join(words)   #去除停用词后组成新的句子
        return sentence


    # 时间分析
    def group_by(self,column):
        workbook1=load_workbook('douban.xlsx')
        sheet=workbook1.get_sheet_by_name("Sheet")
        count = 0
        item={}
        for row in sheet.iter_rows():
            if(count == 0):
                count = 1
                continue
            print(row[0].value)
            join_time = row[column].value
            if (column == 4):
                join_time_str = join_time.split(':')[0]
                join_time = int(join_time_str)
            if(join_time in item):
                item[join_time] = item[join_time]+1
            else:
                item[join_time] = 1
        x = []
        y = []
        for i in sorted (item) : 
            if(column == 4):
                join_time = str(int(i))+'点至'+str(int(i)+1)+'点'
                x.append(join_time)
            else:
                x.append(i)
            y.append(item[i])
        if(column == 4):
            plt.xlabel('日期')
        else:
            plt.xlabel('时刻')
        plt.ylabel('短评数量')
        print(y)
        plt.plot(x, y)
        plt.xticks(x, x, rotation=30)
        if(column == 4):
            plt.title('短评数量随着时刻的变化关系')
        else:
            plt.title('短评数量随着日期的变化关系')
        plt.rcParams['font.sans-serif'] = 'SimHei'
        plt.rcParams['axes.unicode_minus'] = False
        if(column == 4):
            plt.savefig('group_bytime.png')
        else:
            plt.savefig('group_bydate.png')
        





if __name__ == "__main__":
    spider = doubanSpider()
    # 数据采集调用这个
    # spider.start_spider()

    # 词云分析，调用这个
    # 0:全部;1:好评;2:中评;3:差评
    word_type = 3
    # spider.generWord(word_type)


    # 时间分析,调用这个
    # 3:按照日期;4按照时刻
    column = 3
    spider.group_by(column)

