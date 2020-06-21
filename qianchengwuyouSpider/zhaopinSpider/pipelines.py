# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

import os
from openpyxl import Workbook,load_workbook
import pymysql
from zhaopinSpider.config import host,port,user,passwd,db,charset
import time
# 默认的保存方式，存储到excel文件
class ZhaopinspiderPipeline(object):
    path = '51job'+time.strftime('%Y.%m.%d',time.localtime(time.time()))+'.xlsx'
    def __init__(self):
        # 如果已存在，则打开，否则新建
        if(os.path.exists(self.path)):
            self.wb = load_workbook(self.path)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append([ '工作名称','地理位置','学历要求','工作经历','招聘人数', '公司福利','薪资' , '职位描述'
            , '公司名称','公司性质','公司所在行业','公司规模','公司地点', '公司描述','公司招聘主页', '更新时间' , '招聘详细链接'])
 
    def process_item(self, item, spider):
        line = [item['jobName'], item['position'], item['eduLevel'], item['workingExp'], item['peopleNum']
        , item['welfare'], item['salary'] , item['jobDetail'] , item['companyName'], item['companyType']
       , item['companyIndustry'], item['companySize'], item['companyDisplay'], item['companyDesc'], item['companyUrl']
       , item['updateDate'], item['positionURL']]
        self.ws.append(line)
        self.wb.save(self.path)
        return item

# 保存到mysql数据库
class ZhaopinspiderMysqlPipeline(object):

    def __init__(self):
        self.conn = pymysql.connect(host=host, port=port, user=user, passwd=passwd, db=db, charset=charset)
        self.cursor = self.conn.cursor() 

    def process_item(self, item, spider):
        print(item)
        sql = ("INSERT job_info(job_name,position,edu_level,working_exp,people_num,welfare,salary,job_detail"
        ",company_name,company_type,company_industry,company_size,company_display"
        ",company_desc,company_url,update_date,position_url) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)")
        values =[item['jobName'],item['position'],item['eduLevel'],item['workingExp'],item['peopleNum'],item['welfare'],item['salary'],item['jobDetail'],item['companyName'],item['companyType'],item['companyIndustry'],item['companySize']
        ,item['companyDisplay'],item['companyDesc'],item['companyUrl'],item['updateDate'],item['positionURL']]      
        # print(sql)
        self.cursor.execute(sql,values)
        self.conn.commit()
        return item


    
